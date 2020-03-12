# -*- coding: utf-8 -*-
"""
Created on Fri Jun 16 09:51:31 2017

@author: v-stpurc
"""
import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QRadioButton, QVBoxLayout, QCheckBox, QProgressBar,
    QGroupBox, QComboBox, QLineEdit, QPushButton, QMessageBox, QInputDialog, QDialog, QDialogButtonBox, QSlider, QGridLayout, QHBoxLayout, QFileDialog)
from PyQt5.QtGui import *
#from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtCore import *
from PyQt5 import QtCore , QtGui
#from PyQt5.QtCore import QCoreApplication, QThread, pyqtSignal
#Low Current
sys.path.append("C:\\Users\\v-stpur\\Documents\\XBoxScripts\\instrument_Libraries\\")

import visa
import timerThread
import dutInfoDialog
import scopeShotTimer
import nidaqmx
import nidaqmx.system
import ftd2xx as ftd
#import xlsxwriter
import datetime
from dsmcdbgFunctions import DsmcdbgFunctions
import subprocess
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
version_Info = "Python PyQt5 Configuration File Test Automation V0.1"

class MainWindow(QWidget):

    Scope = 0
    Load = 0

    configID = 0
    pcbaSN = 0
    productSN = 0
    scopeType = 0
    probeType = 0 
    runNotes = 0 
    configFileName = 0
    savedSheetName = "None"

    continue_Signal = pyqtSignal(str)
    
    def __init__(self):
        super(MainWindow, self).__init__()
        
        self.FTDIsn = ""
        self.FTDIsn = self.getFTDIserialNumber()
        print self.FTDIsn
        self.logFilePointer = open("CFTAC_Log_" + self.FTDIsn + ".txt", "w")
        self.logFilePointer.write("CFTAC Run Log File\n")
 
        dialog = dutInfoDialog.DutInfoDialog()
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.exec_()
        self.configID, self.probeType, self.productSN, self.runNotes = dialog.returnInfo()
 
        self.configFileName = QFileDialog.getOpenFileName(self, 'Open file', 'C:\\Users\\v-stpur\\Documents\\CFTAC_Python_Port_Apps_MultiChannel',"Excel files (*.xlsx)")
        self.logFilePointer.write("Config File Name: " + str(self.configFileName[0]) + "\n")
        print "Filename is: " + str(self.configFileName[0])

        self.initUI()
        self.openInstruments()
        
        d = ftd.open(0)  
        deviceList = d.getDeviceInfo()
        self.deviceSerialNum = deviceList['serial']
        self.deviceSerialNum = ""
        d.close()

        self.excelFileName = "CFTAC_Output_" + self.deviceSerialNum + datetime.datetime.now().strftime("%Y_%m_%d_%H_%M") + ".xlsx"
        #Create the output file here so we can put the summary page on it
        #then in the timer thread open the config file and create additional tabs on the output file
        self.workBook = Workbook()
        self.SummarySheet = self.workBook.active
        self.SummarySheet.title = "Summary"
       
        headerCellFormat = NamedStyle(name = "Header")
        headerCellFormat.font = Font(bold=True, size = 16)
        self.workBook.add_named_style(headerCellFormat)
        
        self.logFilePointer.write("FTDI Serial Number " + self.FTDIsn + "\n")
        
        self.SummarySheet['A1'].style = headerCellFormat
        self.SummarySheet['A2'].style = headerCellFormat
        self.SummarySheet['A3'].style = headerCellFormat
        self.SummarySheet['A5'].style = headerCellFormat
        self.SummarySheet['C1'].style = headerCellFormat
        self.SummarySheet['C2'].style = headerCellFormat
        self.SummarySheet['E1'].style = headerCellFormat
        
        self.SummarySheet.cell(1, 1, "FTDI S.N.")
        self.SummarySheet.cell(1, 2, self.FTDIsn)
        self.SummarySheet.cell(1, 5, "FW Version")
        dflashSN = self.FTDIsn[:-1]
        self.SummarySheet.cell(1, 4, dflashSN)
        self.SummarySheet.cell(2, 1, "Product S.N.")
        self.SummarySheet.cell(2, 2, self.pcbaSN)
        self.SummarySheet.cell(2, 3, "Date")
        self.SummarySheet.cell(2, 4, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.SummarySheet.cell(3, 1, "Configuration File")
#        self.SummarySheet.cell(3, 2, self.configFileName)
        self.SummarySheet.cell(4, 1, "Probe Type")
        self.SummarySheet.cell(4, 2, self.probeType)
        self.SummarySheet.cell(4, 3, "Run Notes")
        self.SummarySheet.cell(4, 4, self.runNotes)
        
        version = self.getVersion()
        print version[29:41]
        self.SummarySheet.cell(1, 6, version[29:41])

    def initUI(self):
        
        self.setGeometry(300, 300, 600, 200)
        self.setWindowTitle('CFTAC')
        self.setWindowIcon(QIcon('xbox_icon.ico')) 
        
        instrumentChoiceGroupBox  = QGroupBox()
        instrumentChoiceLayout    = QHBoxLayout()


        scopeLabel = QLabel(self)
        scopeLabel.setText("Scope")
        instrumentChoiceLayout.addWidget(scopeLabel)
  
        self.scopeIDN = QLabel(self)
        self.scopeIDN.setText("Scope")
        instrumentChoiceLayout.addWidget(self.scopeIDN)
        
        instrumentChoiceGroupBox.setLayout(instrumentChoiceLayout)

        progressBarGroupBox  = QGroupBox()
        progressBarLayout    = QHBoxLayout()

        self.progressLabel = QLabel(self)
        self.progressLabel.setText("Test Progress")
        self.progressBar = QProgressBar(self)
        progressBarLayout.addWidget(self.progressLabel)
        progressBarLayout.addWidget(self.progressBar)
        
        progressBarGroupBox.setLayout(progressBarLayout)
         
        stepLabelGroupBox  = QGroupBox()
        stepLabelLayout    = QHBoxLayout()
        self.stepLabel = QLabel(self)
        self.stepLabel.setText("Step Label")
        stepLabelLayout.addWidget(self.stepLabel)
        
        self.runLabel = QLabel(self)
        self.runLabel.setText("Run Label")
        stepLabelLayout.addWidget(self.runLabel)
        
        self.appLabel = QLabel(self)
        self.appLabel.setText("App Label")
        stepLabelLayout.addWidget(self.appLabel)
        
        stepLabelGroupBox.setLayout(stepLabelLayout)

        startButtonGroupBox  = QGroupBox()
        startButtonLayout    = QHBoxLayout()
        self.startStopButton = QPushButton('Start Cal', self)
        self.startStopButton.setGeometry(800, 70, 180, 50)

        self.font = QFont()
        self.font.setBold(True)
        self.font.setPointSize(16)
        self.startStopButton.setFont(self.font)
        self.startStopButton.setStyleSheet('QPushButton {background-color: #A3C1DA; color: black;}')
        self.startStopButton.setText("Start Cal")
        self.startStopButton.clicked[bool].connect(self.startStopTest)
        startButtonLayout.addWidget(self.startStopButton)
        startButtonGroupBox.setLayout(startButtonLayout)
 
        quitButtonGroupBox  = QGroupBox()
        quitButtonLayout    = QHBoxLayout()
        self.font.setPointSize(12)
        self.quitButton = QPushButton('Quit', self)
        self.quitButton.setFont(self.font)
        self.quitButton.setGeometry(890, 260, 100, 30)
        self.quitButton.clicked[bool].connect(self.closeEventLocal)
        quitButtonLayout.addWidget(self.quitButton)
        quitButtonGroupBox.setLayout(quitButtonLayout)
        
        grid = QGridLayout()
        grid.setColumnStretch(0,5)
        grid.setColumnStretch(1,5)
        grid.addWidget(instrumentChoiceGroupBox, 0, 0)
        grid.addWidget(progressBarGroupBox, 0, 1)
        grid.addWidget(stepLabelGroupBox, 1, 1)
        grid.addWidget(startButtonGroupBox, 2, 0)
        grid.addWidget(quitButtonGroupBox, 2, 1)
        self.setLayout(grid)

        self.show()

    def closeEventLocal(self, event):
        try:
            self.work.stopTimer()
        except:
            print "Timer Thread Not Running"
            
        self.logFilePointer.close()
        self.logFilePointer = open("CFTAC_Log_" + self.FTDIsn + ".txt", "r")

        #With openpyxl you name the file when you save it.
        #We don't save it until the exit button is pushed and the gui quits
        self.workBook.save(self.excelFileName)

        print("Got quit signal")

        self.logFilePointer.close()
        
        self.close()

    def runLoop(self, numSteps, stepCount, sheetName, run, message, testType):
        
        self.progressBar.setMaximum(numSteps - 1)
        self.progressBar.setMinimum(0)
        self.progressBar.setValue(int(stepCount-1))
        self.stepLabel.setText(sheetName)
        self.runLabel.setText("Run " + str(run))
        self.appLabel.setText(message)

        sys.stdout.write('\a')
        sys.stdout.flush()
        if sheetName != self.savedSheetName:
            QMessageBox.question(self, 'New App Step', 'Move Probes Start App: ' + message, QMessageBox.Ok)
        elif message == "Capture Scope Shot":
            dialog = scopeShotTimer.ScopeShotTimer()
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec_()
        else:
            QMessageBox.question(self, 'New App Step', 'Start App: ' + message, QMessageBox.Ok)
        self.continue_Signal.emit("Continue Test")
        
        self.savedSheetName = sheetName

    def quitLoop(self):
        
        global countGFXruns
        
        self.startStopButton.setStyleSheet('QPushButton {background-color: #A3C1DA; color: black;}')
        self.startStopButton.setText("Start Cal")
        
        self.thread.quit()

    def is_number(self, s):
        try:
            float(s)
            return True
        except ValueError:
            return False   
        
    def startStopTest(self):
        returnString = self.startStopButton.text()
        if returnString.find("Start Cal") != -1:
        
            print "Test Started"

            self.startStopButton.setStyleSheet('QPushButton {background-color: #FF2000; color: black;}')
            self.startStopButton.font()
            self.startStopButton.setText("Abort Cal")
            
            self.startThread()
           
        else:
            print "Test Aborted"
            self.startStopButton.setStyleSheet('QPushButton {background-color: #A3C1DA; color: black;}')
            self.startStopButton.setText("Start Cal")

            self.thread.quit()
            self.work.stopTimer()
#               self.work.startStop = False

    def startThread(self):
        
        self.adInstr = "NI"
        
        parameterArray = [self.Scope, self.logFilePointer, self.configFileName[0], self.configID, self.pcbaSN, self.productSN, self.runNotes, self.workBook, self.SummarySheet]

        self.logFilePointer.write("Starting Thread\n")
        self.work = timerThread.TimerThread(parameterArray) 
        self.work.timerSignal.connect(self.runLoop)
        self.work.quitSignal.connect(self.quitLoop)
        self.continue_Signal.connect(self.work.continue_Slot)
        self.thread = QThread()

        self.work.moveToThread(self.thread)
        self.thread.started.connect(self.work.run)
        self.thread.start()
      
    def openInstruments(self):
        rm = visa.ResourceManager()
        instrumentList = {}
        instrumentList = rm.list_resources()
        
        for loopIndex in range(0, len(instrumentList)):
            
            inst = rm.open_resource(instrumentList[loopIndex])
            try:
                returnString = inst.query("*IDN?")
                print returnString
                if returnString.find("TEKTRONIX,MSO58") != -1:
                    print("found scope")
                    self.Scope = inst
                    self.scopeIDN.setText(returnString[10:23])
            except visa.VisaIOError:
                print "bad juju in instrument list"
        

    def getFTDIserialNumber(self):  
        d = ftd.open(0)  
        deviceList = d.getDeviceInfo()
        self.deviceSerialNum = deviceList['serial']
#        self.deviceSerialNum = self.deviceSerialNum[:0] 
        print "FTDI Serial Number is: " + str(self.deviceSerialNum)
        d.close()
        return str(self.deviceSerialNum)
            
    def flashBoard(self):
        p = subprocess.Popen("C:\Users\v-stpur\Documents\WorkDocs\Tools_6_24\tools>Sflash /remotefile:smcfw.bin /rawwrite:\\gameshare2\ieb\yukon\Toledo\SMC\images\released\SMCFW_21_f108_00\SMCFW_21_f108_00.mng", stdout=subprocess.PIPE, shell=True)
        returnString = str(p.communicate())
        return returnString
            
    def getVersion(self):
        p = subprocess.Popen("dsmcdbg status -i", stdout=subprocess.PIPE, shell=True)
        returnString = str(p.communicate())
        return returnString

if __name__ == '__main__':
    
    app = QCoreApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    ex = MainWindow()
    app.exec_()  
#    sys.exit(app.exec_())  
